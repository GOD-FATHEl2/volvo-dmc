"""
VOLVO DMC Generator - Flask Backend Application
Professional Data Matrix Code generator for VOLVO manufacturing processes

© 2025 VOLVO Cars. All rights reserved. Made by: Nawoar Ekkou
"""

import os
import json
import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
from generate_qr import generate_dmc_code
# from read_dmc import read_dmc_from_bytes  # Original detection
from dmc_detection_hybrid import read_dmc_hybrid  # New hybrid detection
import io

app = Flask(__name__)
CORS(app)

# 🔒 Robust path handling
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "qrs")
LOG_FILE = os.path.join(BASE_DIR, "database.json")
TEMPLATE_FOLDER = os.path.join(BASE_DIR, "templates")
STATIC_FOLDER = os.path.join(BASE_DIR, "static")

app = Flask(__name__, template_folder=TEMPLATE_FOLDER, static_folder=STATIC_FOLDER)
CORS(app)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Setup folders/files safely
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
if not os.path.exists(LOG_FILE):
    with open(LOG_FILE, "w") as f:
        json.dump([], f)


@app.route("/")
def home():
    return render_template("index.html")

@app.route("/login", methods=["POST"])
def login():
    data = request.json
    if data["username"] == "admin" and data["password"] == "admin":
        return jsonify({"success": True})
    return jsonify({"success": False}), 401

@app.route("/generate", methods=["POST"])
def generate():
    prefix = request.form.get("prefix", "").strip().upper()
    count = int(request.form.get("count", 30))

    if not prefix or not prefix.isalnum() or len(prefix) != 1:
        return jsonify({"error": "Invalid prefix"}), 400

    now = datetime.datetime.now()
    date_part = f"{now.month:01d}{now.day:01d}{str(now.year)[-1]}"
    time_part = f"{now.hour:01d}{now.second:01d}"
    dmc_value = f"{prefix}{date_part}{time_part}"  # e.g. A7725640

    with open(LOG_FILE, "r+") as f:
        history = json.load(f)
        existing = {entry["content"] for entry in history}
        results = []
        codes = []
        timestamp = now.strftime("%Y%m%d%H%M%S")

        for i in range(count):
            final_dmc = dmc_value + str(i) if dmc_value in existing else dmc_value
            filename = secure_filename(f"dmc_{i}_{timestamp}.png")
            filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            
            # Generate DMC code and get base64
            base64_image = generate_dmc_code(final_dmc, filepath)

            entry = {
                "content": final_dmc,
                "file": filename,
                "timestamp": str(now)
            }
            history.append(entry)
            results.append(entry)
            
            # Add to codes array for frontend
            codes.append({
                "text": final_dmc,
                "base64": base64_image
            })

        f.seek(0)
        json.dump(history, f, indent=2)
        f.truncate()

    return jsonify({"codes": codes})

@app.route("/history")
def history():
    with open(LOG_FILE, "r") as f:
        return jsonify(json.load(f))

@app.route("/read_dmc", methods=["POST"])
def read_dmc():
    """
    Read DMC code from uploaded image
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    # Check if file is an image
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff'}
    if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
        return jsonify({"error": "Invalid file type. Please upload an image."}), 400
    
    try:
        # Read the image data
        image_data = io.BytesIO(file.read())
        
        # Decode the DMC code
        decoded_text = read_dmc_hybrid(image_data)
        
        if decoded_text:
            # Log the read operation
            now = datetime.datetime.now()
            read_entry = {
                "content": decoded_text,
                "operation": "read",
                "filename": file.filename,
                "timestamp": str(now)
            }
            
            # Add to history
            with open(LOG_FILE, "r+") as f:
                history = json.load(f)
                history.append(read_entry)
                f.seek(0)
                json.dump(history, f, indent=2)
                f.truncate()
            
            return jsonify({
                "success": True,
                "decoded_text": decoded_text,
                "filename": file.filename,
                "timestamp": str(now)
            })
        else:
            return jsonify({"error": "No DMC code found in the image"}), 400
            
    except Exception as e:
        return jsonify({"error": f"Error processing image: {str(e)}"}), 500

@app.route("/read_dmc_camera", methods=["POST"])
def read_dmc_camera():
    """
    Read DMC code from camera capture (base64 image data)
    """
    try:
        # Log incoming request for debugging
        print(f"Camera read request received. Content-Type: {request.content_type}")
        
        data = request.json
        if not data:
            print("Error: No JSON data received")
            return jsonify({"error": "No JSON data provided"}), 400
            
        if 'image' not in data:
            print("Error: No 'image' field in JSON data")
            print(f"Available fields: {list(data.keys())}")
            return jsonify({"error": "No image data provided"}), 400
        
        # Extract base64 image data
        image_data = data['image']
        print(f"Image data length: {len(image_data)} characters")
        
        # Remove data URL prefix if present (data:image/png;base64,)
        if image_data.startswith('data:image'):
            image_data = image_data.split(',')[1]
            print("Removed data URL prefix")
        
        # Decode base64 to bytes
        import base64
        try:
            image_bytes = base64.b64decode(image_data)
            print(f"Successfully decoded base64. Image bytes length: {len(image_bytes)}")
        except Exception as decode_error:
            print(f"Base64 decode error: {decode_error}")
            return jsonify({"error": f"Invalid base64 image data: {str(decode_error)}"}), 400
        
        image_stream = io.BytesIO(image_bytes)
        
        # Decode the DMC code
        decoded_text = read_dmc_hybrid(image_stream)
        print(f"DMC decode result: {decoded_text}")
        
        if decoded_text:
            # Log the read operation
            now = datetime.datetime.now()
            read_entry = {
                "content": decoded_text,
                "operation": "camera_read",
                "filename": "camera_capture",
                "timestamp": str(now)
            }
            
            # Add to history
            with open(LOG_FILE, "r+") as f:
                history = json.load(f)
                history.append(read_entry)
                f.seek(0)
                json.dump(history, f, indent=2)
                f.truncate()
            
            return jsonify({
                "success": True,
                "decoded_text": decoded_text,
                "source": "camera",
                "timestamp": str(now)
            })
        else:
            print("No DMC code found in camera image")
            return jsonify({"error": "No DMC code found in camera image"}), 400
            
    except Exception as e:
        error_msg = f"Error processing camera image: {str(e)}"
        print(f"Camera processing error: {error_msg}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": error_msg}), 500

@app.route("/test_qr")
def test_qr():
    """
    Generate a test QR code for camera testing
    """
    try:
        import qrcode
        import base64
        
        # Generate a simple test QR code using qrcode directly
        test_data = "VOLVO-TEST-12345"
        
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(test_data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to base64
        buffer = io.BytesIO()
        img.save(buffer, 'PNG')
        buffer.seek(0)
        base64_string = base64.b64encode(buffer.getvalue()).decode('utf-8')
        base64_img = f"data:image/png;base64,{base64_string}"
        
        return f"""
        <html>
        <head><title>Test QR Code</title></head>
        <body style="text-align: center; font-family: Arial;">
            <h2>🧪 Test QR Code for Camera</h2>
            <p>Use this QR code to test your camera scanner:</p>
            <div style="border: 2px solid #333; display: inline-block; padding: 20px; margin: 20px;">
                <img src="{base64_img}" style="width: 200px; height: 200px;" />
            </div>
            <p><strong>Data:</strong> {test_data}</p>
            <p>📱 Point your camera at this QR code to test detection</p>
            <a href="/">← Back to VOLVO DMC Generator</a>
        </body>
        </html>
        """
            
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route("/qrs/<path:filename>")
def get_qr(filename):
    return send_from_directory("static/qrs", filename)

@app.route("/download_excel", methods=["POST"])
def download_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.drawing import Image
        
        request_data = request.json or {}
        selected_codes = request_data.get("files", [])
        
        with open(LOG_FILE) as f:
            logs = json.load(f)
        selected = [l for l in logs if l["content"] in selected_codes]

        wb = Workbook()
        ws = wb.active
        ws.title = "DMC Export"
        
        # Add logo to Excel
        logo_path = os.path.join(BASE_DIR, "static", "logo.png")
        if os.path.exists(logo_path):
            try:
                logo_img = Image(logo_path)
                logo_img.width = 80  # Resize logo
                logo_img.height = 40
                ws.add_image(logo_img, 'A1')
                start_row = 4  # Start data below logo
            except:
                start_row = 1
        else:
            start_row = 1
        
        # Add title
        ws.cell(row=start_row, column=1, value="VOLVO DMC Export")
        ws.merge_cells(f'A{start_row}:E{start_row}')
        start_row += 2
        
        # Add codes in 5 columns
        row = []
        current_row = start_row

        for i, log in enumerate(selected):
            row.append(log["content"])
            if len(row) == 5:  # 5 codes per row
                for col, value in enumerate(row, 1):
                    ws.cell(row=current_row, column=col, value=value)
                current_row += 1
                row = []
        
        # Add remaining codes if any
        if row:
            while len(row) < 5:
                row.append("")
            for col, value in enumerate(row, 1):
                ws.cell(row=current_row, column=col, value=value)

        ws.oddFooter.center.text = "© 2025 VOLVO Cars. All rights reserved. Made by: Nawoar Ekkou"
        file = os.path.join(BASE_DIR, "static", f"Export_{datetime.datetime.now().timestamp()}.xlsx")
        wb.save(file)
        return send_file(file, as_attachment=True)
    
    except Exception as e:
        return jsonify({"error": f"Error generating Excel: {str(e)}"}), 500

@app.route("/download_pdf", methods=["POST"])
def download_pdf():
    try:
        from fpdf import FPDF
        request_data = request.json or {}
        selected_codes = request_data.get("files", [])
        
        # Get the file paths from the database
        with open(LOG_FILE) as f:
            logs = json.load(f)
        selected = [l for l in logs if l["content"] in selected_codes]

        class PDF(FPDF):
            def header(self):
                # Add logo
                logo_path = os.path.join(BASE_DIR, "static", "logo.png")
                if os.path.exists(logo_path):
                    try:
                        self.image(logo_path, 10, 8, 33)  # Logo at top left
                    except:
                        pass  # Skip logo if there's an error
                
                self.set_font("Arial", "B", 16)
                self.cell(0, 10, "VOLVO DMC Export", 0, 1, "C")
                self.ln(10)  # Add some space after header

            def footer(self):
                self.set_y(-15)
                self.set_font("Arial", "I", 8)
                self.set_text_color(150)
                self.cell(0, 10, "VOLVO Cars Torslanda All rights reserved. © 2025 Made By: Nawoar Ekkou", 0, 0, "C")

        pdf = PDF()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Add codes to PDF with 5 per row, 5 rows per page (25 total per page)
        for i, log in enumerate(selected):
            if i % 25 == 0:  # New page every 25 codes
                pdf.add_page()
            
            # Calculate position (5 columns, 5 rows)
            col = i % 5
            row = (i // 5) % 5
            
            x = 20 + col * 35  # 35mm spacing between columns
            y = 40 + row * 35  # 35mm spacing between rows
            
            # Add QR code image
            qr_path = os.path.join(app.config["UPLOAD_FOLDER"], log["file"])
            if os.path.exists(qr_path):
                try:
                    pdf.image(qr_path, x=x, y=y, w=25, h=25)
                    # Add text below QR code
                    pdf.set_xy(x, y + 26)
                    pdf.set_font("Arial", "B", 8)
                    pdf.cell(25, 4, log["content"], 0, 0, "C")
                except Exception as e:
                    print(f"Error adding image {qr_path}: {e}")

        pdf_file = os.path.join(BASE_DIR, "static", f"Export_{datetime.datetime.now().timestamp()}.pdf")
        pdf.output(pdf_file)
        return send_file(pdf_file, as_attachment=True)
    
    except Exception as e:
        return jsonify({"error": f"Error generating PDF: {str(e)}"}), 500

if __name__ == "__main__":
    # Get port from environment variable (for Azure Web App) or default to 8000
    port = int(os.environ.get("PORT", 8000))
    app.run(debug=False, host="0.0.0.0", port=port)
