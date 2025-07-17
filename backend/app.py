import os
import json
import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
from generate_qr import generate_dmc_code
from read_dmc import read_dmc_from_bytes
import io

app = Flask(__name__)
CORS(app)

# 🔒 Robust path handling
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "qrs")
LOG_FILE = os.path.join(BASE_DIR, "database.json")

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Setup folders/files safely
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
if not os.path.exists(LOG_FILE):
    with open(LOG_FILE, "w") as f:
        json.dump([], f)


@app.route("/")
def home():
    return render_template("index.html")

@app.route("/login", methods=["POST"])
def login():
    data = request.json
    if data["username"] == "admin" and data["password"] == "admin":
        return jsonify({"success": True})
    return jsonify({"success": False}), 401

@app.route("/generate", methods=["POST"])
def generate():
    prefix = request.form.get("prefix", "").strip().upper()
    count = int(request.form.get("count", 30))

    if not prefix or not prefix.isalnum() or len(prefix) != 1:
        return jsonify({"error": "Invalid prefix"}), 400

    now = datetime.datetime.now()
    date_part = f"{now.month:01d}{now.day:01d}{str(now.year)[-1]}"
    time_part = f"{now.hour:01d}{now.second:01d}"
    dmc_value = f"{prefix}{date_part}{time_part}"  # e.g. A7725640

    with open(LOG_FILE, "r+") as f:
        history = json.load(f)
        existing = {entry["content"] for entry in history}
        results = []
        timestamp = now.strftime("%Y%m%d%H%M%S")

        for i in range(count):
            final_dmc = dmc_value + str(i) if dmc_value in existing else dmc_value
            filename = secure_filename(f"dmc_{i}_{timestamp}.png")
            filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            generate_dmc_code(final_dmc, filepath)

            entry = {
                "content": final_dmc,
                "file": filename,
                "timestamp": str(now)
            }
            history.append(entry)
            results.append(entry)

        f.seek(0)
        json.dump(history, f, indent=2)
        f.truncate()

    return jsonify(results)

@app.route("/history")
def history():
    with open(LOG_FILE, "r") as f:
        return jsonify(json.load(f))

@app.route("/read_dmc", methods=["POST"])
def read_dmc():
    """
    Read DMC code from uploaded image
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    # Check if file is an image
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff'}
    if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
        return jsonify({"error": "Invalid file type. Please upload an image."}), 400
    
    try:
        # Read the image data
        image_data = io.BytesIO(file.read())
        
        # Decode the DMC code
        decoded_text = read_dmc_from_bytes(image_data)
        
        if decoded_text:
            # Log the read operation
            now = datetime.datetime.now()
            read_entry = {
                "content": decoded_text,
                "operation": "read",
                "filename": file.filename,
                "timestamp": str(now)
            }
            
            # Add to history
            with open(LOG_FILE, "r+") as f:
                history = json.load(f)
                history.append(read_entry)
                f.seek(0)
                json.dump(history, f, indent=2)
                f.truncate()
            
            return jsonify({
                "success": True,
                "decoded_text": decoded_text,
                "filename": file.filename,
                "timestamp": str(now)
            })
        else:
            return jsonify({"error": "No DMC code found in the image"}), 400
            
    except Exception as e:
        return jsonify({"error": f"Error processing image: {str(e)}"}), 500

@app.route("/read_dmc_camera", methods=["POST"])
def read_dmc_camera():
    """
    Read DMC code from camera capture (base64 image data)
    """
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({"error": "No image data provided"}), 400
        
        # Extract base64 image data
        image_data = data['image']
        
        # Remove data URL prefix if present (data:image/png;base64,)
        if image_data.startswith('data:image'):
            image_data = image_data.split(',')[1]
        
        # Decode base64 to bytes
        import base64
        image_bytes = base64.b64decode(image_data)
        image_stream = io.BytesIO(image_bytes)
        
        # Decode the DMC code
        decoded_text = read_dmc_from_bytes(image_stream)
        
        if decoded_text:
            # Log the read operation
            now = datetime.datetime.now()
            read_entry = {
                "content": decoded_text,
                "operation": "camera_read",
                "filename": "camera_capture",
                "timestamp": str(now)
            }
            
            # Add to history
            with open(LOG_FILE, "r+") as f:
                history = json.load(f)
                history.append(read_entry)
                f.seek(0)
                json.dump(history, f, indent=2)
                f.truncate()
            
            return jsonify({
                "success": True,
                "decoded_text": decoded_text,
                "source": "camera",
                "timestamp": str(now)
            })
        else:
            return jsonify({"error": "No DMC code found in camera image"}), 400
            
    except Exception as e:
        return jsonify({"error": f"Error processing camera image: {str(e)}"}), 500

@app.route("/qrs/<path:filename>")
def get_qr(filename):
    return send_from_directory("static/qrs", filename)

@app.route("/download_excel", methods=["POST"])
def download_excel():
    from openpyxl import Workbook
    from openpyxl.drawing import Image
    selected_files = request.json.get("files", [])
    with open(LOG_FILE) as f:
        logs = json.load(f)
    selected = [l for l in logs if l["file"] in selected_files]

    wb = Workbook()
    ws = wb.active
    ws.title = "DMC Export"
    
    # Add logo to Excel
    logo_path = os.path.join(BASE_DIR, "static", "logo.png")
    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path)
            logo_img.width = 80  # Resize logo
            logo_img.height = 40
            ws.add_image(logo_img, 'A1')
            start_row = 4  # Start data below logo
        except:
            start_row = 1
    else:
        start_row = 1
    
    # Add title
    ws.cell(row=start_row, column=1, value="VOLVO DMC Export")
    ws.merge_cells(f'A{start_row}:E{start_row}')
    start_row += 2
    
    row = []
    current_row = start_row

    for i, log in enumerate(selected, 1):
        row.append(log["content"])
        if len(row) == 5:
            for col, value in enumerate(row, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1
            row = []
    if row:
        while len(row) < 5:
            row.append("")
        for col, value in enumerate(row, 1):
            ws.cell(row=current_row, column=col, value=value)

    ws.oddFooter.center.text = "VOLVO Cars Torslanda All rights reserved. © 2025 Made By: Nawoar Ekkou"
    file = os.path.join(BASE_DIR, "static", f"Export_{datetime.datetime.now().timestamp()}.xlsx")
    wb.save(file)
    return send_file(file, as_attachment=True)

@app.route("/download_pdf", methods=["POST"])
def download_pdf():
    from fpdf import FPDF
    selected_files = request.json.get("files", [])
    with open(LOG_FILE) as f:
        logs = json.load(f)
    selected = [l for l in logs if l["file"] in selected_files]

    class PDF(FPDF):
        def header(self):
            # Add logo
            logo_path = os.path.join(BASE_DIR, "static", "logo.png")
            if os.path.exists(logo_path):
                self.image(logo_path, 10, 8, 33)  # Logo at top left
            
            self.set_font("Arial", "B", 16)
            self.cell(0, 10, "VOLVO DMC Export", 0, 1, "C")
            self.ln(10)  # Add some space after header

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "I", 8)
            self.set_text_color(150)
            self.cell(0, 10, "VOLVO Cars Torslanda All rights reserved. © 2025 Made By: Nawoar Ekkou", 0, 0, "C")

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    for i, log in enumerate(selected):
        if i % 30 == 0:
            pdf.add_page()
        x = 10 + (i % 5) * 38
        y = 20 + ((i // 5) % 6) * 38
        pdf.image(os.path.join(app.config["UPLOAD_FOLDER"], log["file"]), x=x, y=y, w=10, h=10)

    pdf_file = os.path.join(BASE_DIR, "static", f"Export_{datetime.datetime.now().timestamp()}.pdf")
    pdf.output(pdf_file)
    return send_file(pdf_file, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)
